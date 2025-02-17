import argparse
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill

def import_nessus_file(input_file):
    """Imports the .nessus file and parses it into a BeautifulSoup object."""
    try:
        with open(input_file, "r", encoding="utf-8") as file:
            soup = BeautifulSoup(file, "xml")
        return soup
    except FileNotFoundError:
        raise FileNotFoundError(f"The file {input_file} does not exist.")
    except Exception as e:
        raise Exception(f"An error occurred while reading the file: {e}")

def process_nessus_file(soup):
    """Processes the parsed Nessus file to extract relevant data into a DataFrame."""
    data_rows = []
    unique_tags = set()

    for host in soup.find_all("ReportHost"):
        for tag in host.find_all(recursive=True):
            unique_tags.add(tag.name)

    unique_tags = sorted(unique_tags)

    for host in soup.find_all("ReportHost"):
        row = {}
        row["ReportHost Name"] = host.get("name", "Unknown")

        for tag_name in unique_tags:
            tag = host.find(tag_name)
            row[tag_name] = tag.text if tag else None

        for cvss_tag in host.find_all("cvss_base_score"):
            parent = cvss_tag.find_parent()
            parent_data = row.copy()

            for child in parent.find_all(recursive=False):
                parent_data[child.name] = child.text

            data_rows.append(parent_data)

    df = pd.DataFrame(data_rows)
    return df

def format_dataframe(df):
    """Formats the DataFrame by renaming, reordering columns, and sorting."""
    df.rename(columns={"ReportHost Name": "Host"}, inplace=True)
    df2=df.loc[:, ['cve','cvss3_base_score','cvss_base_score','exploit_available','risk_factor','description','synopsis','solution','cisa-known-exploited','vendor_severity','see_also','rhsa','plugin_output','age_of_vuln']]
    df2.sort_values(by='cvss3_base_score', ascending=False, inplace=True)
    return df2

def export_dataframe(df, output_file):
    """Exports the DataFrame to an Excel file with styled formatting."""
    try:
        writer = pd.ExcelWriter(output_file, engine="openpyxl")
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        sheet = writer.sheets["Sheet1"]

        header_fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")
        row_fill_1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        row_fill_2 = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        for cell in sheet[1]:
            cell.fill = header_fill

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column)):
            fill = row_fill_1 if row_index % 2 == 0 else row_fill_2
            for cell in row:
                cell.fill = fill

        writer.close()
        print(f"Export complete with styling! Data saved to {output_file}")
    except PermissionError:
        raise PermissionError(f"Permission denied: Unable to write to {output_file}.")
    except Exception as e:
        raise Exception(f"An unexpected error occurred while saving the file: {e}")

def main():
    parser = argparse.ArgumentParser(description="Process a Nessus file and export it to an Excel file.")
    parser.add_argument("input_file", help="Path to the input .nessus file.")
    parser.add_argument("output_file", help="Path to the output Excel file.")
    args = parser.parse_args()
    # Validate file extensions
    if not args.input_file.endswith(".nessus"):
        print("Error: The input file must have a .nessus extension.")
        return

    if not args.output_file.endswith(".xlsx"):
        print("Error: The output file must have a .xlsx extension.")
        return
    try:
        # Import Nessus file
        soup = import_nessus_file(args.input_file)

        # Process Nessus file
        df = process_nessus_file(soup)

        # Format DataFrame
        df = format_dataframe(df)

        # Export DataFrame
        export_dataframe(df, args.output_file)

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    main()
